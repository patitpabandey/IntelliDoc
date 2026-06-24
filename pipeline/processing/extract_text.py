"""
extract_text.py
Snowpark Python stored procedure: EXTRACT_TEXT_PROC()

Reads documents in status STAGED from per-client internal stages,
extracts full text using pdfplumber (PDF), openpyxl (XLSX), or the
built-in csv module (CSV), stores the result into DOCUMENT_TEXT,
then advances status to TEXT_EXTRACTED.

Supported formats: PDF | XLSX | CSV

Deployed to Snowflake via:
    CREATE OR REPLACE PROCEDURE EXTRACT_TEXT_PROC()
    RETURNS VARCHAR
    LANGUAGE PYTHON
    RUNTIME_VERSION = '3.11'
    PACKAGES = ('snowflake-snowpark-python', 'pdfplumber', 'openpyxl')
    HANDLER = 'extract_text.run';
"""

from __future__ import annotations

import csv
import io
import re
from typing import Optional

# These imports are available inside Snowpark Python sandbox
try:
    import pdfplumber
    import openpyxl
    from snowflake.snowpark import Session
    from snowflake.snowpark.files import SnowflakeFile
except ImportError:
    pass  # allows local import without Snowpark installed


# ── Text extraction helpers ───────────────────────────────────────────────────

def extract_pdf_text(file_bytes: bytes) -> str:
    """Extract all text from a PDF byte payload."""
    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=2, y_tolerance=2)
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def extract_xlsx_text(file_bytes: bytes) -> str:
    """Extract all cell values from every sheet as concatenated text."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row).strip()
            if row_text:
                text_parts.append(row_text)
    wb.close()
    return "\n".join(text_parts)


def extract_csv_text(file_bytes: bytes) -> str:
    """
    Parse a CSV file and render it as readable labelled text.

    Each row is emitted as  Header: value  pairs separated by  |  so the LLM
    can read field names alongside values without needing to interpret tabular
    structure.  A short header summary is prepended so classification prompts
    can immediately identify the column schema.

    Handles UTF-8 and UTF-8-with-BOM; falls back to latin-1 for legacy files.
    """
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = file_bytes.decode("latin-1", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []

    lines: list[str] = [f"Columns: {', '.join(headers)}"]
    for i, row in enumerate(reader):
        parts = [f"{k}: {v}" for k, v in row.items() if v is not None and str(v).strip()]
        lines.append(" | ".join(parts))

    return "\n".join(lines)


def extract_text(file_bytes: bytes, file_format: str) -> str:
    """Dispatch to the correct extractor based on file format."""
    fmt = file_format.upper()
    if fmt == "PDF":
        return extract_pdf_text(file_bytes)
    if fmt == "XLSX":
        return extract_xlsx_text(file_bytes)
    if fmt == "CSV":
        return extract_csv_text(file_bytes)
    raise ValueError(f"Unsupported file format: {file_format}")


def clean_text(raw: str) -> str:
    """Normalise whitespace without collapsing paragraph breaks."""
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in raw.splitlines()]
    text = re.sub(r"\n{3,}", "\n\n", "\n".join(lines))
    return text.strip()


# ── Snowpark entry point ──────────────────────────────────────────────────────

_CREATE_TEXT_TABLE = """
CREATE TABLE IF NOT EXISTS DOCUMENT_TEXT (
    FILE_ID        VARCHAR(36)        NOT NULL,
    EXTRACTED_TEXT VARCHAR(1000000)   NOT NULL,
    CHAR_COUNT     NUMBER,
    EXTRACTED_TS   TIMESTAMP_NTZ      NOT NULL DEFAULT SYSDATE(),
    CONSTRAINT PK_DOCUMENT_TEXT PRIMARY KEY (FILE_ID)
)
"""

def run(session: "Session") -> str:
    """Snowpark handler — called by EXTRACT_TASK."""
    session.sql(_CREATE_TEXT_TABLE).collect()

    pending = session.sql("""
        SELECT dr.FILE_ID, dr.FILE_NAME, dr.FILE_FORMAT, dr.CLIENT_ID,
               dr.CURRENT_LOCATION, cm.CLIENT_STAGE_NAME
        FROM   DOCUMENT_REGISTRY dr
        JOIN   CLIENT_MAPPING cm ON dr.CLIENT_ID = cm.CLIENT_ID
        WHERE  dr.PROCESSING_STATUS = 'STAGED'
          AND  NOT EXISTS (
               SELECT 1 FROM DOCUMENT_TEXT dt WHERE dt.FILE_ID = dr.FILE_ID
          )
    """).collect()

    extracted, failed = 0, 0

    for row in pending:
        file_id     = row["FILE_ID"]
        file_name   = row["FILE_NAME"]
        file_format = row["FILE_FORMAT"]
        stage_name  = row["CLIENT_STAGE_NAME"]

        session.sql(f"""
            INSERT INTO DOCUMENT_PROCESSING_AUDIT (FILE_ID, STEP_NAME, STATUS, START_TS)
            VALUES ('{file_id}', 'EXTRACT', 'STARTED', SYSDATE())
        """).collect()

        try:
            stage_path = f"@{stage_name}/{file_name}"
            with SnowflakeFile.open(stage_path, "rb") as f:
                file_bytes = f.read()

            raw_text   = extract_text(file_bytes, file_format)
            clean      = clean_text(raw_text)
            char_count = len(clean)
            safe_text  = clean.replace("'", "''")

            session.sql(f"""
                INSERT INTO DOCUMENT_TEXT (FILE_ID, EXTRACTED_TEXT, CHAR_COUNT)
                SELECT '{file_id}', $1, {char_count}
                FROM VALUES ('{safe_text[:999000]}')
                WHERE NOT EXISTS (SELECT 1 FROM DOCUMENT_TEXT WHERE FILE_ID = '{file_id}')
            """).collect()

            session.sql(f"""
                UPDATE DOCUMENT_REGISTRY
                SET PROCESSING_STATUS = 'TEXT_EXTRACTED', UPDATED_TS = SYSDATE()
                WHERE FILE_ID = '{file_id}'
            """).collect()

            session.sql(f"""
                INSERT INTO DOCUMENT_PROCESSING_AUDIT (FILE_ID, STEP_NAME, STATUS, START_TS, END_TS)
                VALUES ('{file_id}', 'EXTRACT', 'COMPLETED', SYSDATE(), SYSDATE())
            """).collect()

            extracted += 1

        except Exception as err:
            err_msg = str(err)[:4000].replace("'", "''")
            session.sql(f"""
                INSERT INTO DOCUMENT_PROCESSING_AUDIT
                    (FILE_ID, STEP_NAME, STATUS, START_TS, END_TS, ERROR_MESSAGE)
                VALUES ('{file_id}', 'EXTRACT', 'FAILED', SYSDATE(), SYSDATE(), '{err_msg}')
            """).collect()
            session.sql(f"""
                UPDATE DOCUMENT_REGISTRY
                SET PROCESSING_STATUS = 'FAILED', UPDATED_TS = SYSDATE()
                WHERE FILE_ID = '{file_id}'
            """).collect()
            failed += 1

    import json
    return json.dumps({"extracted": extracted, "failed": failed})


# ── Snowflake DDL ─────────────────────────────────────────────────────────────
DEPLOY_SQL = """
CREATE OR REPLACE PROCEDURE EXTRACT_TEXT_PROC()
    RETURNS VARCHAR
    LANGUAGE PYTHON
    RUNTIME_VERSION = '3.11'
    PACKAGES = ('snowflake-snowpark-python', 'pdfplumber', 'openpyxl')
    IMPORTS   = ('@INTELLIDOC_PYTHON_STAGE/extract_text.py')
    HANDLER   = 'extract_text.run'
    EXECUTE AS CALLER;
"""

# ── Local test ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from pathlib import Path
    import sys

    source_files = Path(__file__).parent.parent.parent / "Source Files"
    samples      = Path(__file__).parent.parent.parent / "samples"
    search_dirs  = [d for d in (source_files, samples) if d.exists()]

    if not search_dirs:
        print("No Source Files or samples/ directory found.")
        sys.exit(0)

    for d in search_dirs:
        print(f"\n=== {d.name} ===")
        for pdf_path in sorted(d.glob("*.pdf"))[:3]:
            text = clean_text(extract_pdf_text(pdf_path.read_bytes()))
            print(f"PDF  {pdf_path.name}: {len(text)} chars | {text[:120]!r}")
        for xlsx_path in sorted(d.glob("*.xlsx"))[:2]:
            text = clean_text(extract_xlsx_text(xlsx_path.read_bytes()))
            print(f"XLSX {xlsx_path.name}: {len(text)} chars | {text[:120]!r}")
        for csv_path in sorted(d.glob("*.csv"))[:4]:
            text = clean_text(extract_csv_text(csv_path.read_bytes()))
            print(f"CSV  {csv_path.name}: {len(text)} chars | {text[:120]!r}")
